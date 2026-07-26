import pytest
from cockpit.ingestion.parsers.audit_bom import _split_ref_des

def test_split_ref_des_period_accepted():
    assert _split_ref_des("A1.R1, A1.R2") == ("A1.R1", "A1.R2")

def test_split_ref_des_lowercase_accepted():
    assert _split_ref_des("r1, c2") == ("r1", "c2")

def test_split_ref_des_wrappers_stripped_with_content():
    assert _split_ref_des("R1(2), R3") == ("R1", "R3")
    assert _split_ref_des("{DNP}R5") == ("R5",)
    assert _split_ref_des("[note]R7, R8") == ("R7", "R8")
    assert _split_ref_des("R1*alt*") == ("R1",)

def test_split_ref_des_commas_and_hyphens_inside_wrappers_neutralized():
    assert _split_ref_des("R1(a, b), R2") == ("R1", "R2")
    assert _split_ref_des("R1[sub-part], R2") == ("R1", "R2")

def test_split_ref_des_unbalanced_wrapper_fails():
    with pytest.raises(ValueError, match="INVALID_REFDES_TOKEN"):
        _split_ref_des("R1(no-close")

def test_split_ref_des_nested_wrapper_deterministically_fails():
    with pytest.raises(ValueError, match="INVALID_REFDES_TOKEN"):
        _split_ref_des("R1(a(b))")

def test_split_ref_des_symbol_only_token_passes():
    assert _split_ref_des(".") == (".",)
    assert _split_ref_des("-") == ("-",)

def test_split_ref_des_hyphen_unconstrained():
    assert _split_ref_des("R1-R5") == ("R1-R5",)
    assert _split_ref_des("-R1, R2-") == ("-R1", "R2-")

def test_split_ref_des_plus_accepted():
    assert _split_ref_des("TP_12V+") == ("TP_12V+",)
    assert _split_ref_des("+SYS") == ("+SYS",)

def test_split_ref_des_plus_with_annotation():
    assert _split_ref_des("TP_12V+, *PLEASE X-RAY*, (note +5V)") == ("TP_12V+",)

def test_split_ref_des_rejects_comma():
    from cockpit.ingestion.parsers.audit_bom import REFDES_TOKEN_REGEX
    assert REFDES_TOKEN_REGEX.match(",") is None

def test_split_ref_des_rejects_out_of_set_chars():
    with pytest.raises(ValueError, match="INVALID_REFDES_TOKEN"):
        _split_ref_des("R1#")
    with pytest.raises(ValueError, match="INVALID_REFDES_TOKEN"):
        _split_ref_des("R1$")


import openpyxl
from openpyxl.styles import Font
from cockpit.ingestion.errors import MalformedBomError
from cockpit.ingestion.parsers.audit_bom import parse, REQUIRED_HEADER

def _create_bom_xlsx(path, headers, rows, sheet_name="AUDIT BOM"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(str(path))
    wb.close()

def test_parse_14_col_canonical_layout(tmp_path):
    path = tmp_path / "TEST-123 AUDIT BOM.xlsx"
    headers = [
        "Find#", "PartNum", "Count", "MSL level", "Date code", "Baked date",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    rows = [
        [1, "MPN-001", 10, "1", "2026", "01/01", "R1", "0603", "Resistor", "T", 10, 10, 0, "none"]
    ]
    _create_bom_xlsx(path, headers, rows)
    res = parse(path)
    assert res.header_layout == "canonical"
    assert len(res.items) == 1
    assert res.items[0].component_mpn == "MPN-001"
    assert res.items[0].ref_des_raw == "R1"
    assert res.items[0].description == "Resistor"

def test_parse_11_col_legacy_layout_matches_14_col(tmp_path):
    path_14 = tmp_path / "TEST-14 AUDIT BOM.xlsx"
    headers_14 = [
        "Find#", "PartNum", "Count", "MSL level", "Date code", "Baked date",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    rows_14 = [
        [1, "MPN-001", 10, "1", "2026", "01/01", "R1, R2", "0603", "Resistor", "T", 10, 10, 0, "none"],
        [2, "MPN-002", 5, "", "", "", "C1", "0402", "Capacitor", "S", 5, 5, 0, ""]
    ]
    _create_bom_xlsx(path_14, headers_14, rows_14)
    res_14 = parse(path_14)

    path_11 = tmp_path / "TEST-11 AUDIT BOM.xlsx"
    headers_11 = [
        "Find#", "PartNum", "Count",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    rows_11 = [
        [1, "MPN-001", 10, "R1, R2", "0603", "Resistor", "T", 10, 10, 0, "none"],
        [2, "MPN-002", 5, "C1", "0402", "Capacitor", "S", 5, 5, 0, ""]
    ]
    _create_bom_xlsx(path_11, headers_11, rows_11)
    res_11 = parse(path_11)

    assert res_11.header_layout == "legacy"
    assert res_14.header_layout == "canonical"
    assert res_11.items == res_14.items
    assert res_11.raw_row_count == res_14.raw_row_count

def test_parse_partial_optional_layout_12_and_13_col(tmp_path):
    path_12 = tmp_path / "TEST-12 AUDIT BOM.xlsx"
    headers_12 = [
        "Find#", "PartNum", "Count", "MSL level",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    rows_12 = [[1, "MPN-001", 10, "1", "R1", "0603", "Resistor", "T", 10, 10, 0, "none"]]
    _create_bom_xlsx(path_12, headers_12, rows_12)
    res_12 = parse(path_12)
    assert res_12.header_layout == "legacy"
    assert len(res_12.items) == 1

    path_13 = tmp_path / "TEST-13 AUDIT BOM.xlsx"
    headers_13 = [
        "Find#", "PartNum", "Count", "MSL level", "Date code",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    rows_13 = [[1, "MPN-001", 10, "1", "2026", "R1", "0603", "Resistor", "T", 10, 10, 0, "none"]]
    _create_bom_xlsx(path_13, headers_13, rows_13)
    res_13 = parse(path_13)
    assert res_13.header_layout == "legacy"
    assert len(res_13.items) == 1

def test_reordered_required_column_raises_header_drift(tmp_path):
    path = tmp_path / "TEST-ORD AUDIT BOM.xlsx"
    headers = [
        "PartNum", "Find#", "Count",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    _create_bom_xlsx(path, headers, [[1, "MPN-001", 10, "R1", "0603", "Resistor", "T", 10, 10, 0, "none"]])
    with pytest.raises(MalformedBomError) as exc_info:
        parse(path)
    assert exc_info.value.reason == "HEADER_DRIFT"
    assert exc_info.value.detail["required"] == REQUIRED_HEADER

def test_duplicate_required_header_raises_duplicate_header(tmp_path):
    path = tmp_path / "TEST-DUP AUDIT BOM.xlsx"
    headers = [
        "Find#", "Find#", "Count",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    _create_bom_xlsx(path, headers, [[1, 2, 10, "R1", "0603", "Resistor", "T", 10, 10, 0, "none"]])
    with pytest.raises(MalformedBomError) as exc_info:
        parse(path)
    assert exc_info.value.reason == "DUPLICATE_HEADER"
    assert exc_info.value.detail["label"] == "Find#"

def test_entirely_blank_first_row_raises_empty_header(tmp_path):
    path = tmp_path / "TEST-EMP AUDIT BOM.xlsx"
    headers = ["", "", "", ""]
    _create_bom_xlsx(path, headers, [[1, "MPN-001", 10, "R1"]])
    with pytest.raises(MalformedBomError) as exc_info:
        parse(path)
    assert exc_info.value.reason == "EMPTY_HEADER"

def test_leading_or_interior_blank_header_cell_raises_header_drift(tmp_path):
    path_int = tmp_path / "TEST-INT AUDIT BOM.xlsx"
    headers_int = [
        "Find#", "PartNum", "", "Count",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    _create_bom_xlsx(path_int, headers_int, [[1, "MPN", "", 10, "R1", "0603", "Resistor", "T", 10, 10, 0, "none"]])
    with pytest.raises(MalformedBomError) as exc_info:
        parse(path_int)
    assert exc_info.value.reason == "HEADER_DRIFT"

    path_lead = tmp_path / "TEST-LEAD AUDIT BOM.xlsx"
    headers_lead = [""] + REQUIRED_HEADER
    _create_bom_xlsx(path_lead, headers_lead, [[1, "MPN", 10, "R1", "0603", "Resistor", "T", 10, 10, 0, "none", ""]])
    with pytest.raises(MalformedBomError) as exc_info:
        parse(path_lead)
    assert exc_info.value.reason == "HEADER_DRIFT"

def test_data_row_shorter_than_min_row_width_is_skipped(tmp_path):
    path = tmp_path / "TEST-SHORT AUDIT BOM.xlsx"
    headers = [
        "Find#", "PartNum", "Count",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    rows = [
        [1, "MPN-001", 10, "R1", "0603", "Resistor", "T", 10, 10, 0, "none"],
        [2, "MPN-002", 5],  # Too short! Should be skipped without IndexError.
        [3, "MPN-003", 5, "R3", "0603", "Resistor 3", "T", 5, 5, 0, "none"]
    ]
    _create_bom_xlsx(path, headers, rows)
    res = parse(path)
    assert len(res.items) == 2
    assert res.items[0].component_mpn == "MPN-001"
    assert res.items[1].component_mpn == "MPN-003"
    assert res.raw_row_count == 3

def test_strikethrough_extraction_targets_resolved_indices_in_legacy_layout(tmp_path):
    path = tmp_path / "TEST-STRIKE AUDIT BOM.xlsx"
    headers = [
        "Find#", "PartNum", "Count",
        "Ref_Des", "Package", "Description", "SMT/THT", "Qty Need", "Qty On hand",
        "Qty short", "comment"
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AUDIT BOM"
    ws.append(headers)
    
    row = [1, "MPN-001", 10, "R1, R2", "0603", "Resistor", "T", 10, 10, 0, "none"]
    ws.append(row)
    ws.cell(row=2, column=4).font = Font(strike=True)  # Col D is column=4 (Ref_Des)
    
    wb.save(str(path))
    wb.close()
    
    res = parse(path)
    assert len(res.items) == 1
    assert res.items[0].ref_des_raw == ""  # Struck out in resolved Ref_Des column!
    assert res.items[0].description == "Resistor"  # Unaffected
