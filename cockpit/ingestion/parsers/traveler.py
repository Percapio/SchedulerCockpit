"""Traveler metadata parser."""

import datetime
import pathlib
from typing import Any

import openpyxl

from ..errors import AnchorNotFound, CoercionError, MalformedTravelerError
from .coordinate_map import TravelerCoordinateMap
from .results import TravelerResult


def _normalize_anchor(text: Any) -> str:
    """Normalize text for anchor comparison."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    if s.endswith(":"):
        s = s[:-1]
    return s


def _coerce_value(path: pathlib.Path, field_key: str, value: Any, declared_type: str | None) -> Any:
    """Coerce extracted cell value to declared type."""
    if value is None:
        return None
        
    if declared_type == "string":
        s = str(value).strip()
        return s if s else None
        
    if declared_type == "integer":
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            s = value.strip()
            if not s:
                return None
            try:
                return int(s)
            except ValueError:
                pass
        raise CoercionError(path, field_key, "integer", value)
        
    if declared_type == "date":
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        if isinstance(value, str):
            s = value.strip()
            if not s:
                return None
            try:
                return datetime.date.fromisoformat(s)
            except ValueError:
                pass
            try:
                dt = datetime.datetime.strptime(s, "%m/%d/%Y")
                return dt.date()
            except ValueError:
                pass
        raise CoercionError(path, field_key, "date", value)
        
    return value


def parse(path: pathlib.Path, coord_map: TravelerCoordinateMap) -> TravelerResult:
    """Parse Traveler Excel file and extract metadata."""
    try:
        wb = openpyxl.load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:
        raise MalformedTravelerError(path, "UNREADABLE_WORKBOOK", {"error": str(e)})

    try:
        if coord_map.sheet_name not in wb.sheetnames:
            raise MalformedTravelerError(path, "MISSING_SHEET", {"expected": coord_map.sheet_name, "available": wb.sheetnames})
            
        ws = wb[coord_map.sheet_name]
        
        extracted_fields = {}
        raw_anchor_locations = {}
        
        for anchor in coord_map.anchors:
            try:
                anchor_cell = ws[anchor.anchor_cell]
            except Exception as e:
                if anchor.required:
                    raise AnchorNotFound(path, anchor.field_key, anchor.anchor_text, None, anchor.anchor_cell)
                extracted_fields[anchor.field_key] = None
                continue
                
            cell_val = anchor_cell.value
            observed_norm = _normalize_anchor(cell_val)
            
            expected_texts = anchor.anchor_text if isinstance(anchor.anchor_text, list) else [anchor.anchor_text]
            expected_norms = [_normalize_anchor(t) for t in expected_texts]
            
            if observed_norm not in expected_norms:
                if anchor.required:
                    raise AnchorNotFound(path, anchor.field_key, expected_texts, cell_val, anchor.anchor_cell)
                extracted_fields[anchor.field_key] = None
                continue
                
            raw_anchor_locations[anchor.field_key] = anchor.anchor_cell
            
            target_row = anchor_cell.row + anchor.value_offset[0]
            target_col = anchor_cell.column + anchor.value_offset[1]
            
            try:
                target_cell = ws.cell(row=target_row, column=target_col)
                target_val = target_cell.value
            except Exception:
                target_val = None
                
            coerced_val = _coerce_value(path, anchor.field_key, target_val, anchor.value_type)
            extracted_fields[anchor.field_key] = coerced_val
            
        return TravelerResult(
            sheet_name_used=coord_map.sheet_name,
            extracted_fields=extracted_fields,
            raw_anchor_locations=raw_anchor_locations
        )
    finally:
        wb.close()
