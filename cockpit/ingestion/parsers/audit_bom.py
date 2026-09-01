"""Audit BOM parser."""

import logging
import pathlib
import re
import time
import zipfile
from dataclasses import dataclass

import openpyxl
from typing import Any, Final

logger = logging.getLogger(__name__)

try:
    # openpyxl >= 3.1: word-level formatting survives as rich-text runs
    from openpyxl.cell.rich_text import CellRichText, TextBlock
except ImportError:  # pragma: no cover - legacy openpyxl fallback
    CellRichText = None
    TextBlock = None

from ..errors import MalformedBomError
from ..filename_rules import derive_part_number_from_filename
from .results import BomItem, BomResult, HeaderLayout


_ANNOTATION_RE = re.compile(
    r'\*[^*]*\*'        # asterisk markers      (existing)
    r'|\([^)]*\)'       # parenthesis wrappers  (NEW)
    r'|\{[^}]*\}'       # curly-brace wrappers   (NEW)
    r'|\[[^\]]*\]'      # square-bracket wrappers (NEW)
)

REFDES_TOKEN_REGEX = re.compile(r"^[A-Za-z0-9_.+-]+$")

REQUIRED_HEADER: Final[list[str]] = [
    "Find#", "PartNum", "Count",
    "Ref_Des", "Package", "Description", "SMT/THT",
    "Qty Need", "Qty On hand", "Qty short", "comment",
]

OPTIONAL_HEADER: Final[frozenset[str]] = frozenset(
    {"MSL level", "Date code", "Baked date"}
)


def validate_header(
    observed_labels: list[str],
    path: pathlib.Path,
) -> None:
    if not any(observed_labels):
        raise MalformedBomError(path, "EMPTY_HEADER", {"observed": observed_labels})

    trimmed_observed = list(observed_labels)
    while trimmed_observed and trimmed_observed[-1] == "":
        trimmed_observed.pop()

    valid_labels = set(REQUIRED_HEADER) | OPTIONAL_HEADER
    seen = set()
    for label in trimmed_observed:
        if label in valid_labels:
            if label in seen:
                raise MalformedBomError(path, "DUPLICATE_HEADER", {
                    "observed": trimmed_observed,
                    "label": label,
                })
            seen.add(label)

    remainder = [label for label in trimmed_observed if label not in OPTIONAL_HEADER]
    if remainder != REQUIRED_HEADER:
        raise MalformedBomError(path, "HEADER_DRIFT", {
            "required": REQUIRED_HEADER,
            "optional_ignored": sorted(OPTIONAL_HEADER),
            "observed": trimmed_observed,
        })


def resolve_column_index(observed_labels: list[str], name: str) -> int:
    return observed_labels.index(name)


def normalize_sheet_label(label: str) -> str:
    """Normalize sheet label for case-insensitive matching."""
    return str(label).strip().casefold()


def is_valid_excel_sheet_name(candidate: str) -> bool:
    """Check if the given string could be a valid Excel worksheet name."""
    if not candidate:
        return False
    if len(candidate) > 31:
        return False
    for c in ":\\/?*[]":
        if c in candidate:
            return False
    return True


def choose_bom_sheet(path: pathlib.Path, available_sheet_names: list[str], declared_part_number: str) -> str:
    """Select the worksheet to parse, prioritizing canonical over assembly name."""
    available_norms = {normalize_sheet_label(n): n for n in available_sheet_names}
    
    canon_norm = normalize_sheet_label("AUDIT BOM")
    if canon_norm in available_norms:
        return available_norms[canon_norm]
        
    if not is_valid_excel_sheet_name(declared_part_number):
        raise MalformedBomError(path, "MISSING_SHEET", {
            "expected_canonical": "AUDIT BOM",
            "expected_assembly": None,
            "available": available_sheet_names
        })
        
    decl_norm = normalize_sheet_label(declared_part_number)
    if decl_norm in available_norms:
        return available_norms[decl_norm]
        
    raise MalformedBomError(path, "MISSING_SHEET", {
        "expected_canonical": "AUDIT BOM",
        "expected_assembly": declared_part_number,
        "available": available_sheet_names
    })


def coerce_find_number(raw: Any, path: pathlib.Path, mpn: str) -> int:
    if raw is None or str(raw).strip() == "":
        raise MalformedBomError(path, "MISSING_FIND_NUMBER", {"mpn": mpn})
    try:
        fval = float(raw)
        ival = int(fval)
        if fval != ival or ival < 1:
            raise MalformedBomError(path, "INVALID_FIND_NUMBER", {"mpn": mpn, "raw": raw})
        return ival
    except (ValueError, TypeError):
        raise MalformedBomError(path, "INVALID_FIND_NUMBER", {"mpn": mpn, "raw": raw})


def _workbook_has_strike_runs(path: pathlib.Path) -> bool:
    """Cheap pre-check: does any string carry run-level strikethrough?

    Word-level strikethrough lives in rich-text runs as <rPr><strike/>,
    either in xl/sharedStrings.xml (Excel-authored files) or inline in the
    worksheet XML. openpyxl's read-only mode flattens rich text to plain
    str, so files that contain strike runs must be loaded in normal mode;
    every other file keeps the fast, memory-light read-only path. A raw
    byte scan of the decompressed XML is far cheaper than a full parse.
    """
    try:
        with zipfile.ZipFile(str(path)) as zf:
            names = [n for n in zf.namelist()
                     if n == "xl/sharedStrings.xml"
                     or (n.startswith("xl/worksheets/") and n.endswith(".xml"))]
            for name in names:
                if b"<strike" in zf.read(name):
                    return True
        return False
    except Exception:
        return False


def _cell_font_strike(cell: Any) -> bool:
    """Whole-cell strikethrough, tolerant of read-only/empty cells."""
    try:
        font = cell.font
    except AttributeError:
        return False
    return bool(font is not None and getattr(font, "strike", False))


def _visible_cell_text(cell: Any) -> str | None:
    """Return cell text with struck-through words removed (Phase 32, 2.1).

    Word-level strikethrough arrives as rich-text runs (workbook must be
    loaded with rich_text=True); whole-cell strikethrough arrives via the
    cell font. A fully struck cell yields "" — only the text is dropped,
    the row itself survives.
    """
    value = cell.value
    if value is None:
        return None

    if CellRichText is not None and isinstance(value, CellRichText):
        cell_struck = _cell_font_strike(cell)
        parts = []
        for run in value:
            if TextBlock is not None and isinstance(run, TextBlock):
                font = run.font
                if font is not None and getattr(font, "strike", None):
                    continue
                parts.append(run.text or "")
            else:
                # Unformatted runs inherit the cell-level font.
                if cell_struck:
                    continue
                parts.append(str(run))
        return "".join(parts)

    if _cell_font_strike(cell):
        return ""

    return str(value)


def _split_ref_des(raw: str | None) -> tuple[str, ...]:
    if not raw:
        return ()
    # Strip annotation markers before delimiter inspection so that hyphens
    # inside markers (e.g. "*PLEASE X-RAY*") do not trigger the guard.
    cleaned = _ANNOTATION_RE.sub("", raw).strip()
    if not cleaned:
        return ()

    tokens = [tok.strip() for tok in cleaned.split(",") if tok.strip()]
    for tok in tokens:
        if not REFDES_TOKEN_REGEX.match(tok):
            raise ValueError("INVALID_REFDES_TOKEN")
    return tuple(tokens)


def _open_workbook(path: pathlib.Path) -> openpyxl.Workbook:
    # data_only=True to get values, not formulas.
    # rich_text=True (openpyxl >= 3.1) so word-level strikethrough is
    # visible. Read-only mode flattens rich text, so workbooks that
    # actually contain strike runs are loaded in normal mode instead
    # (detected via a cheap sharedStrings.xml scan).
    load_kwargs = dict(filename=str(path), data_only=True, read_only=True)
    if CellRichText is not None:
        load_kwargs["rich_text"] = True
        if _workbook_has_strike_runs(path):
            load_kwargs["read_only"] = False
    return openpyxl.load_workbook(**load_kwargs)


def parse(path: pathlib.Path) -> BomResult:
    """Parse the Audit BOM Excel file and extract THT items."""
    declared_part_number = derive_part_number_from_filename(path)

    try:
        wb = _open_workbook(path)
    except Exception as e:
        raise MalformedBomError(path, "UNREADABLE_WORKBOOK", {"error": str(e)})

    try:
        selected_sheet_name = choose_bom_sheet(path, wb.sheetnames, declared_part_number)
        ws = wb[selected_sheet_name]
        
        # Check header
        header_row = [cell.value for cell in ws[1]]
        observed_labels = [str(v).strip() if v is not None else "" for v in header_row]
        validate_header(observed_labels, path)

        column_index: dict[str, int] = {
            name: resolve_column_index(observed_labels, name)
            for name in REQUIRED_HEADER
        }
        find_col = column_index["Find#"]
        part_col = column_index["PartNum"]
        ref_des_col = column_index["Ref_Des"]
        description_col = column_index["Description"]
        mount_col = column_index["SMT/THT"]

        min_row_width: int = max(column_index.values()) + 1
        present_optionals = OPTIONAL_HEADER.intersection(observed_labels)
        header_layout: HeaderLayout = "canonical" if len(present_optionals) == len(OPTIONAL_HEADER) else "legacy"

        items = []
        raw_row_count = 0
        excluded_dni_count = 0
        excluded_pcb_count = 0
        
        seen_mpns = set()
        duplicate_mpns = set()
        
        for row_cells in ws.iter_rows(min_row=2):
            raw_row_count += 1

            # Phase 32 (3.1): CPU-bound parsing on a worker thread can starve
            # the UI thread of the GIL on heavy BOMs; yield periodically.
            if raw_row_count % 200 == 0:
                time.sleep(0)

            row = [c.value for c in row_cells]
            if len(row) < min_row_width:
                continue

            flag = row[mount_col]
            if flag is None:
                continue
                
            flag_str = str(flag).strip().upper()
            if flag_str.startswith('T'):
                mount_type = 'T'
            elif flag_str.startswith('S'):
                mount_type = 'S'
            else:
                continue
                
            part_number = str(row[part_col]).strip() if row[part_col] is not None else ""
            if not part_number:
                continue
                
            if part_number.upper() == "DNI":
                excluded_dni_count += 1
                continue
            if part_number.upper() == "PCB":
                excluded_pcb_count += 1
                continue
                
            find_number = coerce_find_number(row[find_col], path, part_number)
                
            # Phase 32 (2.1): strikethrough-aware extraction for Ref_Des and
            # Description; struck words are silently dropped.
            description = _visible_cell_text(row_cells[description_col]) if len(row_cells) > description_col else None
            ref_des_raw = _visible_cell_text(row_cells[ref_des_col]) if len(row_cells) > ref_des_col else None
            
            try:
                ref_des_list = _split_ref_des(ref_des_raw)
            except ValueError as ve:
                raise MalformedBomError(path, str(ve), {"raw": ref_des_raw, "mpn": part_number})
                
            if part_number in seen_mpns:
                duplicate_mpns.add(part_number)
            seen_mpns.add(part_number)
            
            items.append(BomItem(
                find_number=find_number,
                component_mpn=part_number,
                description=description,
                ref_des_raw=ref_des_raw,
                ref_des_list=ref_des_list,
                mount_type=mount_type
            ))
            
        if duplicate_mpns:
            raise MalformedBomError(path, "DUPLICATE_MPN", {"duplicates": list(duplicate_mpns)})
            
        logger.debug("Audit BOM header layout resolved: %s (%s)", header_layout, path.name)
        return BomResult(
            declared_part_number=declared_part_number,
            items=items,
            raw_row_count=raw_row_count,
            excluded_dni_count=excluded_dni_count,
            excluded_pcb_count=excluded_pcb_count,
            header_layout=header_layout,
        )
    finally:
        wb.close()


CANONICAL_COLUMNS: tuple[str, ...] = (
    "Find#", "PartNum", "Count", "Ref_Des", "Package", "Description",
    "SMT/THT", "Qty Need", "Qty On hand", "Qty short", "comment",
    "MSL level", "Date code", "Baked date",
)


@dataclass(frozen=True)
class RawBomRow:
    sheet_ordinal: int          # 1-based physical row index; sheet order
    cells: tuple[str, ...]      # exactly 14, CANONICAL_COLUMNS order, "" absent
    part_number: str            # PartNum, stripped
    description: str            # Description, struck words removed


def read_raw_rows(workbook_path: pathlib.Path) -> list[RawBomRow]:
    try:
        wb = _open_workbook(workbook_path)
    except Exception as e:
        raise MalformedBomError(workbook_path, "UNREADABLE_WORKBOOK", {"error": str(e)})

    try:
        declared_part_number = derive_part_number_from_filename(workbook_path)
        selected_sheet_name = choose_bom_sheet(workbook_path, wb.sheetnames, declared_part_number)
        ws = wb[selected_sheet_name]
        
        header_row = [cell.value for cell in ws[1]]
        observed_labels = [str(v).strip() if v is not None else "" for v in header_row]
        
        # Fails only when neither anchor label is present -- at that point the
        # sheet is not an Audit BOM. A legacy sheet missing one of the two is
        # still worth showing.
        anchors_found = 0
        for anchor in ("Find#", "PartNum"):
            try:
                resolve_column_index(observed_labels, anchor)
                anchors_found += 1
            except ValueError:
                pass
        if anchors_found == 0:
            raise MalformedBomError(workbook_path, "MISSING_REQUIRED_COLUMNS", {"observed": observed_labels})

        column_indices: list[int | None] = []
        for col_name in CANONICAL_COLUMNS:
            try:
                column_indices.append(resolve_column_index(observed_labels, col_name))
            except ValueError:
                column_indices.append(None)
                
        results: list[RawBomRow] = []
        # sheet_ordinal is 1-based physical row index
        sheet_ordinal = 1
        for row_cells in ws.iter_rows(min_row=2):
            sheet_ordinal += 1
            
            cells: list[str] = []
            for col_idx in column_indices:
                if col_idx is None or col_idx >= len(row_cells):
                    cells.append("")
                else:
                    text = _visible_cell_text(row_cells[col_idx])
                    cells.append(text.strip() if text else "")

            if not any(cells):
                continue
                
            part_number = cells[1]
            description = cells[5]
            
            results.append(RawBomRow(
                sheet_ordinal=sheet_ordinal,
                cells=tuple(cells),
                part_number=part_number,
                description=description
            ))
            
        return results
    finally:
        wb.close()
